import openpyxl
import smtplib
from openpyxl.utils import get_column_letter, column_index_from_string
from email.mime.text import MIMEText

worksheet = openpyxl.load_workbook('duesRecords.xlsx')
planilha = worksheet.active
devedores = []

for l in range(2, planilha.max_row+1):
    devedores.append({'nome':planilha[f"A{l}"].value,'email':planilha[f'B{l}'].value,'meses':[]})
    for c in range(1, planilha.max_column+1):
        if c > 2 and planilha[f"{get_column_letter(c)}{l}"].value != 'paid':
            nome = planilha[f"A{l}"].value
            mes = planilha[f"{get_column_letter(c)}1"].value
            for i in devedores:
                if i['nome'] == nome:
                    i['meses'].append(mes)

smtobject = smtplib.SMTP('smtp.gmail.com', 587)
smtobject.starttls()
meuemail = 'test'
senha = 'dfwhlvjwpskbmazm'
smtobject.login(meuemail, senha)

for p in devedores:
    menssagem = f'Olá, {p["nome"]}. Você está recebendo este e-mail pois sua conta possui pagamentos pendentes.\nMeses com pagamento pendentes: '
    if len(p['meses']) > 0:
        for m in p['meses']:
            menssagem = menssagem+f"{m}, "
        print(devedores)
        menssagem = menssagem[0:-2]
        menssagem = menssagem+'.'
        menssagemformatada = MIMEText(menssagem, 'plain', 'utf-8')
        smtobject.sendmail('cobranca@gmail.com', meuemail, f"Subject: Club\n{menssagemformatada}")
