import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
book = openpyxl.load_workbook('tabela.xlsx')
book.create_sheet('tabela')
planilha = book['tabela']

for c in range(1, 11):
    for l in range(1, 12):
        celula = planilha.cell(column=c, row=l)
        celula.value = f"{'' if c == 1 and l == 1 else l*c}"
        
book.save('tabelaex.xlsx')