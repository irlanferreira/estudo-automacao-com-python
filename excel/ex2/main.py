import openpyxl, time
coisas = {'Garlic':0.05, 'Celery':0.02, 'Lemons':0.13}
book = openpyxl.load_workbook('produceSales.xlsx')
planilha = book.active

for c in range(1,3):
    for l in range(1, planilha.max_row+1):
        celula = planilha.cell(column=c, row=l)
        coluna = celula.column
        linhaa = celula.row
        valor = celula.value

        if valor in coisas:
            print(f'Atualizando célula {celula.coordinate}...')
            time.sleep(0.3)

            planilha[f"B{l}"] = planilha[f"B{l}"].value + coisas[f"{valor}"]
book.save('vendas.xlsx')