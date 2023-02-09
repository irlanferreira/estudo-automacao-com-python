import openpyxl

book = openpyxl.load_workbook('example.xlsx')
planilha = book.active
planilha.merge_cells('A10:F23')
book.save('joji.xlsx')
planilha.unmerge_cells('A10:F23')