import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
book = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = book.sheetnames
sheet = book[sheet[0]]
print(sheet)
dados = {}
for i in range(2, sheet.max_row+1):
    cs = sheet[f'A{i}'].value
    state = sheet[f'B{i}'].value
    county = sheet[f'C{i}'].value
    pop = sheet[f'D{i}'].value
