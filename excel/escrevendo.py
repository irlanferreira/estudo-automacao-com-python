import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

book = openpyxl.load_workbook('comida.xlsx')
planilha = book.active
cell = planilha.cell(column=1, row=1)
cell.value = 'nao'
book.save('comida.xlsx')