import openpyxl, os, csv
from openpyxl.utils import column_index_from_string, get_column_letter

for arquivo in os.listdir('.//arquivos'):
    if arquivo.endswith('.xlsx'):
        nomearq = os.path.splitext(arquivo)[0]
        book = openpyxl.load_workbook(f'.//arquivos//{arquivo}')
        for planilhanome in book.sheetnames:
            sheet = book[planilhanome]
            lista = []
            for l in range(1, sheet.max_row+1):
                listadados = []
                for c in range(1, sheet.max_column+1):
                    listadados.append(sheet[f'{get_column_letter(c)}{l}'].value)
                lista.append(listadados)
            outputfile = open(f'.//arquivos//csvs//{nomearq}_CONVERTED.csv', 'w', newline='')
            writeobj = csv.writer(outputfile)
            for linha in lista:

                writeobj.writerow(linha)